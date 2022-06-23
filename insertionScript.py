import xlrd 
import openpyxl
import traceback

#TODO : what i need 
#TODO : def a function to get team id from the division and team name and the groupe *
#TODO : for each , modify from att to id 


#! index starts at 1

#list of teams 

def get_teams():
    wookbook = openpyxl.load_workbook("./Finished/equipes.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i,column=5).value == "Honneur" or ws.cell(row=i,column=5).value == "Pré-Honneur":
            tmp_li.append({
                "idEquipe":ws.cell(row=i,column=1).value,
                "sigle":ws.cell(row=i,column=2).value,
                "groupe":ws.cell(row=i,column=4).value,
                "division":ws.cell(row=i,column=5).value,
            })
    return tmp_li

def get_stades():
    wookbook = openpyxl.load_workbook("./Finished/rencontre.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
        if not ws.cell(row=i,column=8).value in tmp_li:
            tmp_li.append(ws.cell(row=i,column=8).value)
    return tmp_li

def get_arbitres():
    wookbook = openpyxl.load_workbook("./Finished/rencontre.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
        if not ws.cell(row=i,column=13).value in tmp_li:
            tmp_li.append(ws.cell(row=i,column=13).value)
    return tmp_li

def get_joueurs():
    wookbook = openpyxl.load_workbook("./Finished/joueurs.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
       tmp_li.append({"idJoueur":ws.cell(row=i,column=1).value,"nomJoueur":ws.cell(row=i,column=3).value})
    return tmp_li

def get_equipe_id(equipes,sigle):
    for eq in equipes:
        if eq["sigle"]==sigle :
            return eq["idEquipe"]
    raise Exception("Id team not found")

def get_stade_id(stades,nomStade):
    return stades.index(nomStade)

def get_arbitre_id(arbitres,nomArbitre):
    return arbitres.index(nomArbitre)

def get_joueur_id(joueurs,nomJoueur):
    # return 1
    for eq in joueurs:
        if eq["nomJoueur"]==nomJoueur :
            return eq["idJoueur"]
    raise Exception("Id={0} joueur not found".format(nomJoueur))

#adding equipe 
def insert_equipe_script():
    file = open('insert_equipe_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/equipes.xlsx")
    ws = wookbook.active
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i,column=5).value == "Honneur" or ws.cell(row=i,column=5).value == "Pré-Honneur":
            file.write("INSERT INTO Equipe(idEquipe,sigle,nomComplet,division,groupe,anneeFondation,telephone,mobile,fax,location) VALUES ({a},'{b}','{c}','{d}','{m}',{e},{f},{g},{h},{i});\n".format(
                a=ws.cell(row=i,column=1).value,
                b=ws.cell(row=i,column=2).value.replace("'"," "),
                c=ws.cell(row=i,column=3).value.replace("'"," "),
                m=ws.cell(row=i,column=4).value,
                d=ws.cell(row=i,column=5).value,
                e=ws.cell(row=i,column=6).value or 'NULL',
                f=ws.cell(row=i,column=7).value or 'NULL',
                g=ws.cell(row=i,column=8).value or 'NULL', 
                h=ws.cell(row=i,column=9).value or 'NULL',
                i= "'{0}'".format(ws.cell(row=i,column=13).value.replace("'"," ")) if ws.cell(row=i,column=13).value != None else 'NULL'
            ))
    file.close()

#adding staff

def insert_staff_script(equipes):
    file1 = open('insert_staff_script.sql', 'a',encoding="utf-8")
    file2 = open('insert_occuper_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/staff.xlsx")
    ws = wookbook.active
    for i in range(2, ws.max_row+1):
        try:
            file1.write("INSERT INTO Staff(idStaff,prenom,nom,dateNaissance,lieuNaissance,wilayaNaissance) VALUES ({a},'{b}','{c}','{d}','{e}','{f}');\n".format(
                a=ws.cell(row=i,column=1).value,
                b=ws.cell(row=i,column=3).value.split(" ")[0].replace("'"," "),
                c=ws.cell(row=i,column=3).value.split(" ")[1].replace("'"," "),
                d=ws.cell(row=i,column=6).value,
                e=ws.cell(row=i,column=7).value.replace("'"," ") or 'NULL',
                f=ws.cell(row=i,column=8).value.replace("'"," ") or 'NULL',
            ))
            file2.write("INSERT INTO Occuper(idStaff,idEquipe,idSeason,position) VALUES ({a},{b},'2021/2022','{d}');\n".format(
                a=ws.cell(row=i,column=1).value,
                b=get_equipe_id(equipes,ws.cell(row=i,column=4).value),
                d=ws.cell(row=i,column=2).value.replace("'"," "),
            ))
        except:
            pass
        # print(ws.cell(row=i,column=1).value,"   -",ws.cell(row=i,column=4).value,"-")
        # print(ws.cell(row=i,column=1).value,"   -",get_equipe_id(equipes,ws.cell(row=i,column=4).value))
    file1.close()
    file2.close()


#adding entraineur 

def insert_entraineur_script(equipes):
    file1 = open('insert_entraineur_script.sql', 'a',encoding="utf-8")
    file2 = open('insert_entrainer_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/entraineurs.xlsx")
    ws = wookbook.active
    for i in range(2, ws.max_row+1):
        try:
            file1.write("INSERT INTO Entraineur(idEntraineur,nom,prenom,dateNaissance,lieuNaissance,wilayaNaissance) VALUES ({a},'{b}','{c}','{d}','{e}','{f}');\n".format(
                a=ws.cell(row=i,column=1).value,
                b=ws.cell(row=i,column=3).value.split(" ")[0].replace("'"," "),
                c=ws.cell(row=i,column=3).value.split(" ")[1].replace("'"," "),
                d=ws.cell(row=i,column=7).value,
                e=ws.cell(row=i,column=8).value.replace("'"," ") or 'NULL',
                f=ws.cell(row=i,column=9).value.replace("'"," ") or 'NULL',
            ))
            file2.write("INSERT INTO Entrainer(idEntraineur,idEquipe,idSeason,categorie,poste) VALUES ({a},{b},'2021/2022','{d}','{e}');\n".format(
                a=ws.cell(row=i,column=1).value,
                b=get_equipe_id(equipes,ws.cell(row=i,column=4).value),
                d=ws.cell(row=i,column=5).value.replace("'"," "),
                e=ws.cell(row=i,column=2).value.replace("'"," ")
            ))
        except:
            pass
        # print(ws.cell(row=i,column=1).value,"   -",ws.cell(row=i,column=4).value,"-")
        # print(ws.cell(row=i,column=1).value,"   -",get_equipe_id(equipes,ws.cell(row=i,column=4).value))
    file1.close()
    file2.close()


def insert_joueur_script(equipes):
    file1 = open('insert_joueur_script.sql', 'a',encoding="utf-8")
    file2 = open('insert_jouerPour_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/joueurs.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
        try:
            team = get_equipe_id(equipes,ws.cell(row=i,column=5).value)
            file1.write("INSERT INTO Joueur(idJoueur,nom,prenom,dateNaissance,lieuNaissance,wilayaNaissance) VALUES ({a},'{b}','{c}','{d}','{e}','{f}');\n".format(
                a=ws.cell(row=i,column=1).value,
                b=ws.cell(row=i,column=3).value.split(" ")[0].replace("'"," "),
                c=ws.cell(row=i,column=3).value.split(" ")[1].replace("'"," "),
                d=ws.cell(row=i,column=8).value,
                e=ws.cell(row=i,column=9).value.replace("'"," ") or 'NULL',
                f=ws.cell(row=i,column=10).value.replace("'"," ") or 'NULL'
            ))
            file2.write("INSERT INTO JouerPour(idJoueur,idEquipe,idSeason,position,categorie,dossard) VALUES ({a},{b},'2021/2022','{d}','{e}',{f});\n".format(
                a=ws.cell(row=i,column=1).value,
                b=get_equipe_id(equipes,ws.cell(row=i,column=5).value),
                d=ws.cell(row=i,column=2).value.replace("'"," "),
                e=ws.cell(row=i,column=6).value.replace("'"," "),
                f=ws.cell(row=i,column=4).value if ws.cell(row=i,column=4).value != None else 'NULL'
            ))
            tmp_li.append({"idJoueur":ws.cell(row=i,column=1).value,"nomJoueur":ws.cell(row=i,column=3).value})
        except:
            pass
    file1.close()
    file2.close()
    return tmp_li


def insert_rencontre_script(equipes,stades,arbitres):
    file1 = open('insert_rencontre_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/rencontre.xlsx")
    ws = wookbook.active
    tmp_li = []
    for i in range(2, ws.max_row+1):
    # for i in range(2, 5):
        try:
            file1.write("""INSERT INTO Rencontre(idRencontre,idSeason,idStade,categorie,groupe,journee,date,idEquipeA,idEquipeB,butsA,butsB,idArbitre,assisstant1,assisstant2,commissaire,staffMedical,matchJoue)   
            VALUES ({a},'2021/2022',{c},'{d}','{e}',{f},'{g}',{ha},{hb},{ia},{ib},{j},{k},{l},{m},{n},{o});\n""".format(
                a=ws.cell(row=i,column=1).value,
                c= "{0}".format(get_stade_id(stades,ws.cell(row=i,column=8).value)) if ws.cell(row=i,column=8).value != None else 'NULL',
                d=ws.cell(row=i,column=2).value,
                e=ws.cell(row=i,column=3).value,
                f=ws.cell(row=i,column=5).value,
                g=ws.cell(row=i,column=6).value,
                ha=get_equipe_id(equipes,ws.cell(row=i,column=9).value),
                hb=get_equipe_id(equipes,ws.cell(row=i,column=10).value),
                ia = ws.cell(row=i,column=11).value or 0,
                ib = ws.cell(row=i,column=12).value or 0, 
                j = get_arbitre_id(arbitres,ws.cell(row=i,column=13).value),
                k = "'{0}'".format(ws.cell(row=i,column=14).value.replace("'"," ")) if ws.cell(row=i,column=14).value != None else 'NULL',
                l =  "'{0}'".format(ws.cell(row=i,column=15).value.replace("'"," ")) if ws.cell(row=i,column=14).value != None else 'NULL',
                m =  "'{0}'".format(ws.cell(row=i,column=16).value.replace("'"," ")) if ws.cell(row=i,column=14).value != None else 'NULL',
                n =  "'{0}'".format(ws.cell(row=i,column=17).value.replace("'"," ")) if ws.cell(row=i,column=14).value != None else 'NULL',
                o = ws.cell(row=i,column=11).value != None,
            ))
            tmp_li.append(ws.cell(row=i,column=1).value)
        except Exception as ex:
            # traceback.print_exc()
            # print(i,'****',ex)
            pass

    file1.close()
    return tmp_li


def insert_jouer_script(liste_rencontres,joueurs):
    file1 = open('insert_jouer_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/match_joueurs.xlsx")
    ws = wookbook.active
    tmp_li = []
    tmp  = 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i,column=1).value in liste_rencontres:
            try:
                file1.write("""INSERT INTO Jouer(idRencontre,idJoueur,titulaire,capitaine,nombreCartonJaune,nombreCartonRouge,dossard)   
                VALUES ({a},{c},{d},{e},{f},{g},{h});\n""".format(
                    a=ws.cell(row=i,column=1).value,
                    c= get_joueur_id(joueurs,ws.cell(row=i,column=9).value),
                    d=ws.cell(row=i,column=2).value,
                    e=ws.cell(row=i,column=3).value,
                    f="1" if ws.cell(row=i,column=4).value == "TRUE" else "0",
                    g="1" if ws.cell(row=i,column=5).value == "TRUE" else "0",
                    h = ws.cell(row=i,column=8).value,
                ))
                tmp_li.append(ws.cell(row=i,column=1).value)
                # print(i)
            except Exception as ex:
                tmp += 1
                # traceback.print_exc()
                # print("Joueur not found {0}".format(i))
                # print(i,'****',ex)
                pass
        # print(ws.cell(row=i,column=1).value,"   -",ws.cell(row=i,column=4).value,"-")
        # print(ws.cell(row=i,column=1).value,"   -",get_equipe_id(equipes,ws.cell(row=i,column=4).value))
    file1.close()
    print(tmp)
    return tmp_li


def insert_marquer_script(equipes,joueurs,liste_rencontres):
    file1 = open('insert_marquer_script.sql', 'a',encoding="utf-8")
    wookbook = openpyxl.load_workbook("./Finished/buts.xlsx")
    ws = wookbook.active
    tmp= 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i,column=1).value in liste_rencontres :
            try:
                idJoueur = get_joueur_id(joueurs,ws.cell(row=i,column=4).value)
                file1.write("""INSERT INTO Marquer(idRencontre,idJoueur,estPenalite,minute,pourEquipe)   
                VALUES ({a},{c},{d},{e},{f});\n""".format(
                    a=ws.cell(row=i,column=1).value,
                    c= idJoueur,
                    d=ws.cell(row=i,column=2).value,
                    e=ws.cell(row=i,column=3).value,
                    f=get_equipe_id(equipes,ws.cell(row=i,column=5).value),
                ))
            except Exception as ex:
                # traceback.print_exc()
                # print(i,'****',ex)
                tmp +=1
                pass
        # print(ws.cell(row=i,column=1).value,"   -",ws.cell(row=i,column=4).value,"-")
        # print(ws.cell(row=i,column=1).value,"   -",get_equipe_id(equipes,ws.cell(row=i,column=4).value))
    file1.close()
    print(tmp)


def insert_stades_script(stades):
    file1 = open('insert_stade_script.sql', 'a',encoding="utf-8")
    for i in range(0,len(stades)):
        file1.write("""INSERT INTO Stade(idStade,nomStade) VALUES ({a},'{c}');\n""".format(a=i,c= stades[i]))
    file1.close()


def insert_arbitres_script(arbitres):
    file1 = open('insert_arbitre_script.sql', 'a',encoding="utf-8")
    for i in range(0,len(arbitres)):
        file1.write("""INSERT INTO Arbitre(idArbitre,nomArbitre) VALUES ({a},'{c}');\n""".format(a=i,c= arbitres[i]))
    file1.close()



tmp = get_teams()
tmp2 = get_stades()
tmp3 = get_arbitres()
tmp4 = insert_rencontre_script(tmp,tmp2,tmp3) #liste id des rencontres
tmp5 = get_joueurs()

tmp5 = insert_joueur_script(tmp)


tmp6 = insert_jouer_script(tmp4,tmp5) #liste joueurs 
insert_marquer_script(tmp,tmp5,tmp4)


#? Call for functions 
#!--tested
# insert_equipe_script()
# insert_staff_script(tmp)
# insert_entraineur_script(tmp)
# insert_stades_script(tmp2)
# insert_arbitres_script(tmp3)
# tmp4 = insert_rencontre_script(tmp,tmp2,tmp3)
# tmp6 = insert_jouer_script(tmp4,tmp5) #liste joueurs 


#!--not tested 

# tmp6 = insert_jouer_script(tmp4,tmp5)
# insert_marquer_script(tmp,tmp5,tmp4)



# Import openyxl module

# Define variable to load the wookbook
# wookbook = openpyxl.load_workbook("equipes.xlsx")

# Define variable to read the active sheet:
# ws = wookbook.active

# Iterate the loop to read the cell values
# for i in range(0, ws.max_row):
#     for col in ws.iter_cols(1, ws.max_column):
#         print(col[i].value, end="\t\t")
#     print('')

# print(ws.cell(row=1,column=1).value)


# print(get_equipe_id(tmp,"J S B A B","Honneur","Groupe 1"))



