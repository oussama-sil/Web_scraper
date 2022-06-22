from tkinter import E
import requests
from bs4 import BeautifulSoup
import dateparser
import xlsxwriter
import traceback
import openpyxl


def get_digits(str):
    tmp_str=""
    for s in str:
        if s.isdigit():
            tmp_str = tmp_str + s
    return tmp_str

def get_saison(str):
    tmp_str=""
    for s in str:
        if s.isdigit() or s=='/':
            tmp_str = tmp_str + s
    return tmp_str




def get_rencontre_stats(idRencontre):
    URL = "https://www.lwf-alger.org/resultat/view?id="+str(idRencontre)
    page = requests.get(URL,headers={
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    })

    soup = BeautifulSoup(page.content, "html.parser")

    #! Test if id exists 
    if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0].text or "Erreur (#8)" in soup.find_all("div",{"class"  : "wrapper"})[0].text:
        # print("Id:"+str(idRencontre)+" non valide")
        raise Exception("Id:"+str(idRencontre)+" non valide")

    results =   soup.find_all("h3",{"class"  : "result-header"})
    #? Generale infos
    gen_info = soup.find_all("h3",{"class"  : "result-header"})
    tmp_li1 = gen_info[0].text.replace("\r","").split("\n")
    #? finding the categorie groupe honneur 
    tmp_li2 = tmp_li1[1].split("/")
    categorie = tmp_li2[0].strip()
    groupe = tmp_li2[1].strip()
    honneur = tmp_li2[2].strip()
    #? finding journée
    journee = get_digits(tmp_li1[2])

    #?finding date
    date = str(dateparser.parse(tmp_li1[3].strip()).date())

    # print(tmp_li1)

    #?finding season , stade
    tmp_li3 = soup.find_all("div",{"class"  : "result-location"})
    tmp_li4 = tmp_li3[0].findChildren("ul" , recursive=False)[0].findChildren("li" , recursive=False)
    season = get_saison(tmp_li4[0].text.strip())
    stade = tmp_li4[2].text.strip()




    #? Getting the teams 
    tmp_li5 = soup.find_all("span",{"class"  : "d-none d-sm-block"})
    EquipeA = tmp_li5[0].text.strip()
    EquipeB = tmp_li5[1].text.strip()

    #? getting the result of the match 
    tmp_li6 = soup.find_all("div",{"class"  : "result-match"}) 
    butsA = get_digits(tmp_li6[0].text.split(":")[0])
    butsB = get_digits(tmp_li6[0].text.split(":")[1])


    #! the list of goals 
    list_buts = []

    #?buts equipeA
    tmp_li7 = soup.find_all("div",{"class"  : "team"})
    tmp_li8 = tmp_li7[0].findChildren("li" , recursive=True)
    for tmp_but in tmp_li8:
        estPenalite = tmp_but.text != None and "(P)" in tmp_but.text
        minute = get_digits(tmp_but.text)
        if estPenalite:
            nomJoueur = tmp_but.text.split("(")[0].strip()
        else:
            nomJoueur = tmp_but.text.split(minute)[0].strip()
        pour = EquipeA
        list_buts.append({"idRencontre":idRencontre, "estPenalite":estPenalite,"minute":int(minute),"nomJoueur":nomJoueur,"pour":pour})
    #? buts equipeB
    tmp_li7 = soup.find_all("div",{"class"  : "team right"})
    tmp_li8 = tmp_li7[0].findChildren("li" , recursive=True)
    for tmp_but in tmp_li8:
        estPenalite = tmp_but.text != None and "(P)" in tmp_but.text
        minute = get_digits(tmp_but.text)
        if estPenalite:
            nomJoueur = tmp_but.text.split("(")[0].strip()
        else:
            nomJoueur = tmp_but.text.split(minute)[0].strip()
        pour = EquipeB
        list_buts.append({"idRencontre":idRencontre, "estPenalite":estPenalite,"minute":int(minute),"nomJoueur":nomJoueur,"pour":pour})
        
    # print(list_buts)

    #? Les cartons  

    tmp_li9 = soup.find_all("div",{"class"  : "container"})
    tmp_li10 = tmp_li9[5].findChildren("span" ,{"class":"left"}, recursive=True)
    tmp_li11 = tmp_li9[5].findChildren("span" ,{"class":"right"}, recursive=True)

    # tmp_li11 = tmp_li10[0].findChildren("div" , recursive=False)

    list_joueurs_rencontre = []
    #equipeA
    indx = 0
    for tmp_j in tmp_li10:
        if tmp_j.text.strip() != "":
            titulaire = indx <= 10
            capitaine  = "(C)" in str(tmp_j)
            cartonJaune = "carton" in str(tmp_j) and "jaune.png" in str(tmp_j)
            cartonRouge = "carton" in str(tmp_j) and "rouge.png" in str(tmp_j)
            guardien = "(P)" in str(tmp_j)
            numLicence = tmp_j.findChildren("div" , recursive=False)[0].text.strip()
            indx += 1
            numero = tmp_j.findChildren("div" , recursive=False)[1].text.strip()
            if numero == "" :
                numero = "0"
            nom = tmp_j.findChildren("div" , recursive=False)[2].findChildren("span" , recursive=False)[0].text+" "+tmp_j.findChildren("div" , recursive=False)[2].findChildren("span" , recursive=False)[1].text
            list_joueurs_rencontre.append({
                "idRencontre":idRencontre,
                "titulaire":titulaire,
                "capitaine":capitaine,
                "cartonJaune":cartonJaune,
                "cartonRouge":cartonRouge,
                "guardien":guardien,
                "numLicence":numLicence,
                "numero":numero,
                "nom":nom
            })
    #equipeB
    indx = 0
    for tmp_j in tmp_li11:
        if tmp_j.text.strip() != "":
            titulaire = indx <= 10
            capitaine  = "(C)" in str(tmp_j)
            cartonJaune = "carton" in str(tmp_j) and "jaune.png" in str(tmp_j)
            cartonRouge = "carton" in str(tmp_j) and "rouge.png" in str(tmp_j)
            guardien = "(P)" in str(tmp_j)
            numLicence = tmp_j.findChildren("div" , recursive=False)[2].text.strip()
            indx += 1
            numero = tmp_j.findChildren("div" , recursive=False)[1].text.strip()
            if numero == "" :
                numero = None
            nom = tmp_j.findChildren("div" , recursive=False)[0].findChildren("span" , recursive=False)[0].text.strip()+" "+tmp_j.findChildren("div" , recursive=False)[0].findChildren("span" , recursive=False)[1].text.strip()
            list_joueurs_rencontre.append({
                "idRencontre":idRencontre,
                "titulaire":titulaire,
                "capitaine":capitaine,
                "cartonJaune":cartonJaune,
                "cartonRouge":cartonRouge,
                "guardien":guardien,
                "numLicence":numLicence,
                "numero":numero,
                "nom":nom
            })


    #?officiels 

    tmp_li12 = soup.find_all("div",{"class"  : "container officiels"})[1].findChildren("div" , recursive=False)[0].findChildren("div" , recursive=False)

    try:
        arbitre = tmp_li12[0].text.split(":")[1].strip()
    except:
        arbitre = None
    
    try:
        assistant1 = tmp_li12[1].text.split(":")[1].strip()
    except:
        assistant1 = None
    
    try:
        assistant2 = tmp_li12[2].text.split(":")[1].strip()
    except:
        assistant2 = None
    
    try:
        commissaire = tmp_li12[3].text.split(":")[1].strip()
    except:
        commissaire = None
        
    try:
        staffMedical = tmp_li12[4].text.split(":")[1].strip()
    except:
        staffMedical = "-"


    #?l'objet de la rencontre 

    rencontre = {
        "idRencontre":idRencontre,
        "categorie":categorie,
        "groupe":groupe,
        "honneur":honneur,
        "journee":journee,
        "date":date,
        "season":season,
        "stade":stade,
        "EquipeA":EquipeA,
        "EquipeB":EquipeB,
        "butsA":butsA,
        "butsB":butsB,
        "arbitre":arbitre,
        "assistant1":assistant1,
        "assistant2":assistant2,
        "commissaire":commissaire,
        "staffMedical":staffMedical
    }
    return rencontre,list_buts,list_joueurs_rencontre,rencontre["season"]=="2021/2022"



def get_equipe(idEquipe):
    URL = "https://www.lwf-alger.org/club/view?id="+str(idEquipe)
    page = requests.get(URL,headers={
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    })

    soup = BeautifulSoup(page.content, "html.parser")

    #! Test if id exists 
    if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0].text or "Erreur (#8)" in soup.find_all("div",{"class"  : "wrapper"})[0].text:
        # print("Id:"+str(idRencontre)+" non valide")
        raise Exception("Id:"+str(idEquipe)+" non valide")

    sigle = soup.find_all("div",{"class"  : "section-title-team"})[0].findChildren("h1" , recursive=True)[0].text.strip()
    tmp_li1 = soup.find_all("div",{"class"  : "section-title-team"})[0].findChildren("h6" , recursive=True)
    nomComplet = tmp_li1[0].text.split(":")[1].strip()
    groupe = tmp_li1[1].text.split(":")[1].strip()
    division = tmp_li1[2].text.split(":")[1].strip()
    anneeFondation = tmp_li1[3].text.split(":")[1].strip()
    try:
        telephone = int(tmp_li1[4].text.split(":")[1].strip()) 
    except:
        telephone = None
    try:
        mobile = int(tmp_li1[5].text.split(":")[1].strip()) 
    except:
        mobile = None
    try:
        fax =   int(tmp_li1[6].text.split(":")[1].strip()) 
    except:
        fax = None
    
    president = tmp_li1[7].text.split(":")[1].strip()
    presidentDeSection= tmp_li1[8].text.split(":")[1].strip()
    try:
        totalJoueur = int( tmp_li1[9].text.split(":")[1].strip())
    except:
        totalJoueur = None
    location = tmp_li1[10].text.split(":")[1].strip()
    stade = tmp_li1[11].text.split(":")[1].strip()
    
    return {"idEquipe":idEquipe, "sigle":sigle,"nomComplet":nomComplet,"groupe":groupe,
        "division":division,"anneeFondation":anneeFondation,"telephone":telephone,
        "mobile":mobile,"fax":fax,"president":president,"presidentDeSection":presidentDeSection,"totalJoueur":totalJoueur,
        "location":location,"stade":stade    
    }

def create_equipe_file(start,end):
    
    book=xlsxwriter.Workbook("equipes.xlsx")
    sheet=book.add_worksheet()
    row = 0
    col = 0
        #writing the headers 
    headers = ["idEquipe","sigle","nomComplet","groupe",
        "division","anneeFondation","telephone",
        "mobile","fax","president","presidentDeSection","totalJoueur",
        "location","stade"]
    for hd in headers:
        sheet.write(row,col,hd)
        col +=1
        #writing items 
    row +=1
    for i in range(start,end+1):
        col = 0
        try:
            print(i)
            eq = get_equipe(i)
            for key in eq:
                sheet.write(row,col,eq[key])
                col += 1
            row += 1
        except Exception as ex :
            traceback.print_exc()
            print(ex)

        
    book.close()
    


def get_joueur(idJoueur):
    URL = "https://www.lwf-alger.org/joueur/view?id="+str(idJoueur)
    page = requests.get(URL,headers={
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    })

    soup = BeautifulSoup(page.content, "html.parser")

    #! Test if id exists 
    if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0].text or "Erreur (#8)" in soup.find_all("div",{"class"  : "wrapper"})[0].text:
        # print("Id:"+str(idRencontre)+" non valide")
        raise Exception("Id:"+str(idJoueur)+" non valide")
    try:
        numeroJoueur = int(soup.find_all("span",{"class"  : "number-player"})[0].text)
    except:
        numeroJoueur = None
    
    post = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].findChildren("span" , recursive=False)[0].text.strip()
    nomJoueur = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].text.strip()
    nomJoueur = nomJoueur.replace(post,"").strip()
    clubJoueur = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[0].text.replace("Club","").strip()
    categorieJoueur =  soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[1].text.replace("Catégorie","").strip()
    try:
        ageJoueur = int(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[2].text.replace("Age","").strip())
    except:
        ageJoueur = None
    try:
        dateNaissance =  str(dateparser.parse(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[3].text.replace("Date de naissance","").strip()).date())
    except:
        dateNaissance = None
    lieuNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[4].text.replace("Lieu de naissance","").strip()
    wilayaNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[5].text.replace("Wilaya","").strip()
    
    joueur = {
        "idJoueur":idJoueur,
        "post":post,
        "nomJoueur":nomJoueur,
        "numeroJoueur":numeroJoueur,
        "clubJoueur":clubJoueur,
        "categorieJoueur":categorieJoueur,
        "ageJoueur":ageJoueur,
        "dateNaissance":dateNaissance,
        "lieuNaissance":lieuNaissance,
        "wilayaNaissance":wilayaNaissance
    }   

    return joueur,soup.find_all("td")[0].text.strip()=="2021/2022"

def create_joueur_file(start,end,filename):
    book=xlsxwriter.Workbook(filename)
    sheet=book.add_worksheet()
    row = 0
    col = 0
        #writing the headers 
    headers = ["idJoueur",
        "post",
        "nomJoueur",
        "numeroJoueur",
        "clubJoueur",
        "categorieJoueur",
        "ageJoueur",
        "dateNaissance",
        "lieuNaissance",
        "wilayaNaissance"]
    for hd in headers:
        sheet.write(row,col,hd)
        col +=1
        #writing items 
    row +=1
    for i in range(start,end+1):
        col = 0
        try:
            eq,cnd = get_joueur(i)
            print(i)
            if cnd:
                for key in eq:
                    sheet.write(row,col,eq[key])
                    col += 1
                row += 1
        except Exception as ex :
            # traceback.print_exc()
            print(ex)

    book.close()



def get_entraineur(idEntraineur):
    URL = "https://www.lwf-alger.org/entraineur/view?id="+str(idEntraineur)
    page = requests.get(URL,headers={
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    })

    soup = BeautifulSoup(page.content, "html.parser")

    #! Test if id exists 
    if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0].text or "Erreur (#8)" in soup.find_all("div",{"class"  : "wrapper"})[0].text:
        # print("Id:"+str(idRencontre)+" non valide")
        raise Exception("Id:"+str(idEntraineur)+" non valide")
    
    post = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].findChildren("span" , recursive=False)[0].text.strip()
    nom = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].text.strip()
    nom = nom.replace(post,"").strip()
    club = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[0].text.replace("Club","").strip()
    categorie =  soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[1].text.replace("Catégorie","").strip()
    try:
        age = int(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[2].text.replace("Age","").strip())
    except:
        age = 0
    try:
        dateNaissance =  str(dateparser.parse(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[3].text.replace("Date de naissance","").strip()).date())
    except:
        dateNaissance = None
    lieuNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[4].text.replace("Lieu de naissance","").strip()
    wilayaNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[5].text.replace("Wilaya","").strip()
    entraineur = {
        "idEntraineur":idEntraineur,
        "post":post,
        "nom":nom,
        "club":club,
        "categorie":categorie,
        "age":age,
        "dateNaissance":dateNaissance,
        "lieuNaissance":lieuNaissance,
        "wilayaNaissance":wilayaNaissance
    }   

    return entraineur,soup.find_all("td")[0].text.strip()=="2021/2022"

def create_entraineur_file(start,end):
    book=xlsxwriter.Workbook("entraineurs.xlsx")
    sheet=book.add_worksheet()
    row = 0
    col = 0
        #writing the headers 
    headers = ["idEntraineur",
        "post",
        "nom",
        "club",
        "categorie",
        "age",
        "dateNaissance",
        "lieuNaissance",
        "wilayaNaissance"]
    for hd in headers:
        sheet.write(row,col,hd)
        col +=1
        #writing items 
    row +=1
    for i in range(start,end+1):
        col = 0
        try:
            print(i)
            eq,cnd = get_entraineur(i)
            if cnd:
                for key in eq:
                    sheet.write(row,col,eq[key])
                    col += 1
                row += 1
        except Exception as ex :
            traceback.print_exc()
            print(ex)

    book.close()



def get_staff(idStaff):
    URL = "https://www.lwf-alger.org/staff/view?id="+str(idStaff)
    page = requests.get(URL,headers={
    "User-Agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    })

    soup = BeautifulSoup(page.content, "html.parser")

    #! Test if id exists 
    if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0].text or "Erreur (#8)" in soup.find_all("div",{"class"  : "wrapper"})[0].text:
        # print("Id:"+str(idRencontre)+" non valide")
        raise Exception("Id:"+str(idStaff)+" non valide")
    
    post = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].findChildren("span" , recursive=False)[0].text.strip()
    nom = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("h4" , recursive=False)[0].text.strip()
    nom = nom.replace(post,"").strip()
    club = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[0].text.replace("Club","").strip()
    try:
        age = int(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[1].text.replace("Age","").strip())
    except:
        age = 0
    try:
        dateNaissance =  str(dateparser.parse(soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[2].text.replace("Date de naissance","").strip()).date())
    except:
        dateNaissance = None
    lieuNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[3].text.replace("Lieu de naissance","").strip()
    wilayaNaissance = soup.find_all("div",{"class"  : "info-player"})[0].findChildren("li" , recursive=True)[4].text.replace("Wilaya","").strip()
    staff = {
        "idStaff":idStaff,
        "post":post,
        "nom":nom,
        "club":club,
        "age":age,
        "dateNaissance":dateNaissance,
        "lieuNaissance":lieuNaissance,
        "wilayaNaissance":wilayaNaissance
    }   

    return staff,soup.find_all("td")[0].text.strip()=="2021/2022"

def create_staff_file(start,end):
    book=xlsxwriter.Workbook("staff.xlsx")
    sheet=book.add_worksheet()
    row = 0
    col = 0
    #writing the headers 
    headers = ["idStaff",
        "post",
        "nom",
        "club",
        "age",
        "dateNaissance",
        "lieuNaissance",
        "wilayaNaissance"]
    for hd in headers:
        sheet.write(row,col,hd)
        col +=1
        #writing items 
    row +=1
    for i in range(start,end+1):
        col = 0
        try:
            print(i)
            eq,cnd = get_staff(i)
            if cnd:
                for key in eq:
                    sheet.write(row,col,eq[key])
                    col += 1
                row += 1
        except Exception as ex :
            # traceback.print_exc()
            print(ex)

    book.close()

# create_equipe_kfile(1,500)

# create_staff_file(1,1400)

# a,b=get_entraineur(1300)
# print(b)
# create_entraineur_file(1,1400)

# get_joueur(28628)
# create_joueur_file(28620,28628)

# eq = get_equipe(207)
# for key in eq:

#     print(key+":"+str(eq[key]),end="\n")

#! To get data 

idRencontre = 9870
#! 1
# start = 1
# end = 10000
#! 2
# start = 10001
# end = 20000
#! 3
# start = 20001
# end = 30000
#! 4
start = 30001
end = 40000
#! 5
# start = 40001
# end = 45000


#!-----For player
# filename = "joueurC1.xlsx"
# end = 1000
# start = 0
# for i in range(0,50):
#     create_joueur_file(1000*i+1,1000*(i+1),"./joueurs/joueurC{0}.xlsx".format(i))
#     print("")

# #Writing to the file 
def create_rencontre_file(start,end):
    book1=xlsxwriter.Workbook("rencontre.xlsx")
    sheet1=book1.add_worksheet()
    book2=xlsxwriter.Workbook("buts.xlsx")
    sheet2=book2.add_worksheet()
    book3=xlsxwriter.Workbook("match_joueurs.xlsx")
    sheet3=book3.add_worksheet()
    row3 = 0
    col3 = 0
    #writing the headers 
    headers3 = ["idRencontre",
                    "titulaire",
                    "capitaine",
                    "cartonJaune",
                    "cartonRouge",
                    "guardien",
                    "numLicence",
                    "numero",
                    "nom"]
    for hd in headers3:
        sheet3.write(row3,col3,hd)
        col3 +=1
    row3 +=1

    row2 = 0
    col2 = 0
    headers2 = ["idRencontre","estPenalite","minute","nomJoueur","pour"]
    for hd in headers2:
        sheet2.write(row2,col2,hd)
        col2 +=1
    row2+=1

    row1 = 0
    col1 = 0
    headers1 = ["idRencontre",
            "categorie",
            "groupe",
            "honneur",
            "journee",
            "date",
            "season",
            "stade",
            "EquipeA",
            "EquipeB",
            "butsA",
            "butsB",
            "arbitre",
            "assistant1",
            "assistant2",
            "commissaire",
            "staffMedical"]
    for hd in headers1:
        sheet1.write(row1,col1,hd)
        col1 +=1
    row1 +=1


    for i in range(start,end+1):
        col3 = 0
        col2 = 0
        col1 = 0
        try:
            rencontre,list_buts,list_joueurs_rencontre,cnd = get_rencontre_stats(i)

            if cnd : #if in 2021/2022
                print(i)

                for key in rencontre:
                    sheet1.write(row1,col1,rencontre[key])
                    col1 += 1
                row1 += 1
                
                for but in list_buts:
                    col2 = 0
                    for key in but:
                        sheet2.write(row2,col2,but[key])
                        col2 += 1
                    row2 += 1

                for jr in list_joueurs_rencontre:
                    col3 = 0
                    for key in jr:
                        sheet3.write(row3,col3,jr[key])
                        col3 += 1
                    row3 += 1
                
        except Exception as ex :
            # traceback.print_exc()
            print(ex)

        
    book1.close()
    book2.close()
    book3.close()
#!---


def merge_joueurs_files(start,end):
    book=xlsxwriter.Workbook("joueur.xlsx")
    sheet=book.add_worksheet()
    row = 0
    col = 0
    headers = ["idJoueur",
        "post",
        "nomJoueur",
        "numeroJoueur",
        "clubJoueur",
        "categorieJoueur",
        "ageJoueur",
        "dateNaissance",
        "lieuNaissance",
        "wilayaNaissance"]
    for hd in headers:
        sheet.write(row,col,hd)
        col +=1
    row +=1
    col = 0
    for fn in range(start,end+1):
        wookbook = openpyxl.load_workbook("./joueurs/joueurC{0}.xlsx".format(fn))
        ws = wookbook.active
        for i in range(2, ws.max_row+1):   #row of the file ws
            for col in range(1,len(headers)+1): #column  
                sheet.write(row,col-1,ws.cell(row=i,column=col).value)
            row+=1
    book.close()  

merge_joueurs_files(0,10)


# create_rencontre_file(8761,9900)


# book=xlsxwriter.Workbook("buts.xlsx")
# sheet=book.add_worksheet()
# row = 0
# col = 0
#     #writing the headers 
# headers = ["idRencontre","estPenalite","minute","nomJoueur","pour"]
# for hd in headers:
#     sheet.write(row,col,hd)
#     col +=1
#     #writing items 
# row +=1
# for i in range(start,end+1):
#     col = 0
#     try:
#         print(i)
#         rencontre,list_buts,list_joueurs_rencontre,cnd = get_rencontre_stats(i)
            
#         # for key in rencontre:
#         #     print(key+":"+str(rencontre[key]),end="\n")
#         if cnd :
#             for but in list_buts:
#                 col = 0
#                 for key in but:
#                     sheet.write(row,col,but[key])
#                     col += 1
#                 row += 1
#     except Exception as ex :
#         traceback.print_exc()
#         print(ex)

    
# book.close()

# !----
# start = 1
# end = 10000
# book=xlsxwriter.Workbook("rencontre.xlsx")
# sheet=book.add_worksheet()
# row = 0
# col = 0
#     #writing the headers 
# headers = ["idRencontre",
#         "categorie",
#         "groupe",
#         "honneur",
#         "journee",
#         "date",
#         "season",
#         "stade",
#         "EquipeA",
#         "EquipeB",
#         "butsA",
#         "butsB",
#         "arbitre",
#         "assistant1",
#         "assistant2",
#         "commissaire",
#         "staffMedical"]
# for hd in headers:
#     sheet.write(row,col,hd)
#     col +=1
#     #writing items 
# row +=1
# for i in range(start,end+1):
#     col = 0
#     try:
#         print(i)
#         rencontre,list_buts,list_joueurs_rencontre,cnd = get_rencontre_stats(i)
            
#         # for key in rencontre:
#         #     print(key+":"+str(rencontre[key]),end="\n")
#         if cnd:
#           for key in rencontre:
#               sheet.write(row,col,rencontre[key])
#               col += 1
#           row += 1
#     except Exception as ex :
#         traceback.print_exc()
#         print(ex)
# book.close()


#!-----
# if "Not Found (#404)" in soup.find_all("div",{"class"  : "wrapper"})[0]
#     return 0
# print( "carton" in str(tmp_li10[2]) and "jaune.png" in str(tmp_li10[2]))

# print(tmp_li10[17].text)
# print("(C)" in str(tmp_li10[0]))

# pour un joueurs dans un buts : 
# titulaire 
# capitaine
# carton jaune 
# carton rouge 
# minuteEntré
# minuteSortie
#guardien
#numLicence












#print(results.prettify()) #? to print html code 

#getting the list 
# job_elements = results.find_all("div", class_="card-content")
# print(job_elements[0]) #? it s a list 

# for job_element in job_elements:
#     # print(job_element, end="\n"*2)
#     title_element = job_element.find("h2", class_="title")
#     company_element = job_element.find("h3", class_="company")
#     location_element = job_element.find("p", class_="location")
#     print(title_element.text.strip())
#     print(company_element.text.strip())
#     print(location_element.text.strip())
#     print()